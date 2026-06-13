from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def load_jsonish(text: Any) -> Any:
    if text is None:
        return None
    raw = str(text).strip()
    if not raw:
        return None
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # Some hand-written expected cells are almost JSON but miss braces/commas.
        return raw


def normalize_trade(trade: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        str(trade.get("account", "") or ""),
        str(trade.get("amount", "") or ""),
        str(trade.get("term", "") or ""),
        str(trade.get("price", "") or ""),
        str(trade.get("intent", "") or ""),
    )


def compare_files(expected_path: str | Path, actual_path: str | Path) -> list[str]:
    expected_wb = load_workbook(expected_path, read_only=True, data_only=True)
    actual_wb = load_workbook(actual_path, read_only=True, data_only=True)
    expected_ws = expected_wb.active
    actual_ws = actual_wb.active

    expected_headers = [cell.value for cell in next(expected_ws.iter_rows(min_row=1, max_row=1))]
    actual_headers = [cell.value for cell in next(actual_ws.iter_rows(min_row=1, max_row=1))]
    exp_idx = expected_headers.index("预期输出")
    act_idx = actual_headers.index("预期格式输出")

    diffs: list[str] = []
    for idx, (exp_row, act_row) in enumerate(
        zip(
            expected_ws.iter_rows(min_row=2, values_only=True),
            actual_ws.iter_rows(min_row=2, values_only=True),
        ),
        start=2,
    ):
        expected = load_jsonish(exp_row[exp_idx])
        actual = load_jsonish(act_row[act_idx])
        if not isinstance(expected, dict) or not isinstance(actual, dict):
            continue
        exp_trades = [normalize_trade(t) for t in expected.get("trades", [])]
        act_trades = [normalize_trade(t) for t in actual.get("trades", [])]
        if exp_trades != act_trades:
            diffs.append(f"row {idx}: expected={exp_trades} actual={act_trades}")
    expected_wb.close()
    actual_wb.close()
    return diffs

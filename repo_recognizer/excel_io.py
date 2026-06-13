from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from .models import Message, ProcessedRow, clean_text, normalize_bool


REQUIRED_COLUMNS = ["con_ID", "SENDER", "CONTEXT"]


def header_map(headers: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, header in enumerate(headers):
        name = clean_text(header)
        if name:
            result[name] = idx
    return result


def get_value(row: tuple[Any, ...], columns: dict[str, int], name: str) -> Any:
    idx = columns.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def load_messages(path: str | Path) -> tuple[list[str], list[tuple[Any, ...]], list[Message]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean_text(cell) for cell in header_row]
    columns = header_map(headers)
    missing = [name for name in REQUIRED_COLUMNS if name not in columns]
    if missing:
        raise ValueError(f"Missing required columns in {path}: {missing}")

    rows: list[tuple[Any, ...]] = []
    messages: list[Message] = []
    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        rows.append(row)
        raw = {header: get_value(row, columns, header) for header in headers if header}
        messages.append(
            Message(
                row_number=excel_row_number,
                con_id=clean_text(get_value(row, columns, "con_ID")),
                sender=clean_text(get_value(row, columns, "SENDER")),
                context=clean_text(get_value(row, columns, "CONTEXT")),
                send_time=clean_text(get_value(row, columns, "CHATSENDTIMEORI")),
                tradername=clean_text(get_value(row, columns, "TRADERNAME")),
                interlocutor=clean_text(get_value(row, columns, "INTERLOCUTOR")),
                is_start=normalize_bool(get_value(row, columns, "is_start")),
                raw=raw,
            )
        )
    workbook.close()
    return headers, rows, messages


def json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def write_results(
    output_path: str | Path,
    input_headers: list[str],
    input_rows: list[tuple[Any, ...]],
    processed: list[ProcessedRow],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "输出"

    output_headers = list(input_headers) + [
        "识别LLM输出json",
        "判断LLM输出json",
        "llm_used",
        "llm_error",
        "最终状态json",
        "预期格式输出",
    ]
    worksheet.append(output_headers)

    for row, item in zip(input_rows, processed):
        worksheet.append(
            list(row)
            + [
                json_text(item.extract_result),
                json_text(item.judge_result),
                "Y" if item.used_llm else "N",
                item.llm_error,
                json_text(item.final_state),
                json_text(item.public_result),
            ]
        )

    for col_cells in worksheet.columns:
        header = str(col_cells[0].value or "")
        if header in {
            "CONTEXT",
            "识别LLM输出json",
            "判断LLM输出json",
            "llm_error",
            "最终状态json",
            "预期格式输出",
            "预期输出",
        }:
            width = 60
        else:
            width = min(max(len(header) + 2, 12), 24)
        worksheet.column_dimensions[col_cells[0].column_letter].width = width
        for cell in col_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"
    workbook.save(output_path)
    workbook.close()

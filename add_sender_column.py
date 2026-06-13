"""
Add SENDER column after con_ID in multiple xlsx files.
- Header: "SENDER" (uppercase, bold, matching existing header font)
- When SENTTYPE == "发送" (send): SENDER = TRADERNAME
- When SENTTYPE == "接收" (receive): SENDER = INTERLOCUTOR
- Preserves existing column widths (does not change other columns' widths)
"""

import openpyxl
from openpyxl.styles import Font
from copy import copy
import os

FILES = [
    '交易下文_输入与预期输出和实际输出.xlsx',
    '交易下文_测试集.xlsx',
    '交易下文_测试集v1.0输出.xlsx',
    '交易下文_测试集v2.0输出.xlsx',
]


def col_letter_to_index(letter):
    """Convert column letter to 1-based index. A=1, B=2, ..."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def col_index_to_letter(index):
    """Convert 1-based index to column letter. 1=A, 2=B, ..."""
    result = ''
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result


def process_workbook(filepath):
    print(f'\n{"="*60}')
    print(f'Processing: {filepath}')

    wb = openpyxl.load_workbook(filepath)
    modified = False

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'  Sheet: {sheet_name}')

        # --- Find column positions ---
        con_id_col = None
        trader_col = None
        interlocutor_col = None
        senttype_col = None

        for col in range(1, ws.max_column + 1):
            header_val = str(ws.cell(row=1, column=col).value or '').strip()
            if header_val == 'con_ID':
                con_id_col = col
            elif header_val == 'TRADERNAME':
                trader_col = col
            elif header_val == 'INTERLOCUTOR':
                interlocutor_col = col
            elif header_val == 'SENTTYPE':
                senttype_col = col

        # Check if SENDER already exists
        for col in range(1, ws.max_column + 1):
            header_val = str(ws.cell(row=1, column=col).value or '').strip()
            if header_val == 'SENDER':
                print(f'    SENDER already exists at col {col}, skipping.')
                con_id_col = None  # signal to skip
                break

        if con_id_col is None:
            print(f'    Skipping — no con_ID or SENDER already present.')
            continue

        if None in (trader_col, interlocutor_col, senttype_col):
            print(f'    ERROR: missing required columns. con_ID={con_id_col}, TRADERNAME={trader_col}, INTERLOCUTOR={interlocutor_col}, SENTTYPE={senttype_col}')
            continue

        insert_pos = con_id_col + 1  # Insert right after con_ID

        print(f'    con_ID at col {con_id_col}, inserting SENDER at col {insert_pos}')
        print(f'    TRADERNAME at col {trader_col}, INTERLOCUTOR at col {interlocutor_col}, SENTTYPE at col {senttype_col}')

        # --- Save column widths before insert ---
        # openpyxl doesn't shift column_dimensions correctly on insert_cols,
        # so we save them and restore manually.
        saved_widths = {}  # 1-based index -> width
        for col_letter, dim in ws.column_dimensions.items():
            idx = col_letter_to_index(col_letter)
            if dim.width is not None:
                saved_widths[idx] = dim.width

        # --- Get reference font from existing header ---
        ref_cell = ws.cell(row=1, column=con_id_col)
        new_font = copy(ref_cell.font)
        new_font.bold = True

        # --- Insert the new column ---
        ws.insert_cols(insert_pos)

        # --- Restore column widths for shifted columns ---
        # After insert, columns at positions >= insert_pos shifted right by 1.
        # openpyxl shifts the ColumnDimension objects but keeps old letter keys,
        # so dimensions end up on wrong columns. We fix by re-assigning widths
        # to the correct column letters.
        for old_idx, width in saved_widths.items():
            if old_idx >= insert_pos:
                new_idx = old_idx + 1
                new_letter = col_index_to_letter(new_idx)
                ws.column_dimensions[new_letter].width = width
            else:
                # Column didn't shift — keep its width
                old_letter = col_index_to_letter(old_idx)
                ws.column_dimensions[old_letter].width = width

        # Ensure the new column (at insert_pos) does NOT have a custom width override
        # from the shift — remove any customWidth that might have been set
        new_letter = col_index_to_letter(insert_pos)
        if new_letter in ws.column_dimensions:
            # The inserted column may have inherited width from surrounding columns
            # Reset it — let Excel auto-size
            del ws.column_dimensions[new_letter]

        # --- Set SENDER header ---
        header_cell = ws.cell(row=1, column=insert_pos)
        header_cell.value = 'SENDER'
        header_cell.font = new_font

        # --- After insert, find the shifted column positions ---
        # Everything at or after insert_pos shifted right by 1
        new_trader_col = trader_col + 1 if trader_col >= insert_pos else trader_col
        new_interlocutor_col = interlocutor_col + 1 if interlocutor_col >= insert_pos else interlocutor_col
        new_senttype_col = senttype_col + 1 if senttype_col >= insert_pos else senttype_col

        # --- Populate SENDER values ---
        count_send = 0
        count_recv = 0
        errors = 0

        for row in range(2, ws.max_row + 1):
            senttype = str(ws.cell(row=row, column=new_senttype_col).value or '').strip()
            trader = ws.cell(row=row, column=new_trader_col).value
            interlocutor = ws.cell(row=row, column=new_interlocutor_col).value

            if senttype == '发送':
                ws.cell(row=row, column=insert_pos).value = trader
                count_send += 1
            elif senttype == '接收':
                ws.cell(row=row, column=insert_pos).value = interlocutor
                count_recv += 1
            else:
                errors += 1
                print(f'    WARNING: Row {row} has unexpected SENTTYPE: {repr(senttype)}')

        print(f'    Done: {count_send} 发送 rows, {count_recv} 接收 rows')
        if errors:
            print(f'    Errors: {errors}')

        modified = True

    if modified:
        wb.save(filepath)
        print(f'  Saved: {filepath}')
    else:
        print(f'  No changes made.')

    wb.close()


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for filename in FILES:
        filepath = os.path.join(base_dir, filename)
        if os.path.exists(filepath):
            process_workbook(filepath)
        else:
            print(f'FILE NOT FOUND: {filepath}')

    print('\nDone! All files processed.')


if __name__ == '__main__':
    main()
